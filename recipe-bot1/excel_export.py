from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

import db

EXPORT_DIR = Path(__file__).parent / "exports"

HEADERS = [
    "ID",
    "Telegram ID",
    "Имя",
    "Фамилия",
    "Год рождения",
    "Месяц рождения",
    "Увлечения / работа",
    "Создан",
    "Обновлён",
]


def export_users_to_excel() -> Path:
    EXPORT_DIR.mkdir(exist_ok=True)
    filename = f"anketa_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = EXPORT_DIR / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Анкеты"
    ws.append(HEADERS)

    for user in db.get_all_users():
        ws.append([
            user["id"],
            user["user_id"],
            user["first_name"],
            user["last_name"],
            user["birth_year"],
            user["birth_month"],
            user["hobbies"],
            user["created_at"],
            user["updated_at"],
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = [5, 12, 15, 15, 12, 15, 25, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    wb.save(path)
    return path
